from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, List


class ExcelParser:
    
    def parse(self, file_path: str) -> Dict:
        wb = load_workbook(file_path, data_only=True)
        
        sheets_data = []
        full_text = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            rows = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    rows.append([str(cell) if cell is not None else "" for cell in row])
            
            sheet_text = []
            for row in rows:
                row_text = " | ".join(row)
                if row_text.strip():
                    sheet_text.append(row_text)
            
            sheets_data.append({
                "sheet_name": sheet_name,
                "rows": rows,
                "row_count": len(rows),
                "column_count": sheet.max_column
            })
            
            full_text.extend(sheet_text)
        
        wb.close()
        
        return {
            "file_name": Path(file_path).name,
            "file_type": "xlsx",
            "total_sheets": len(sheets_data),
            "full_text": "\n".join(full_text),
            "sheets": sheets_data,
            "metadata": {
                "sheet_names": wb.sheetnames
            }
        }
    
    def extract_sheet(self, file_path: str, sheet_name: str = None) -> List[List[str]]:
        wb = load_workbook(file_path, data_only=True)
        
        if sheet_name:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
        
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append([str(cell) if cell is not None else "" for cell in row])
        
        wb.close()
        return rows
